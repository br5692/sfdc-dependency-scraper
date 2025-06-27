"""
Robust Salesforce Field Dependency Scraper - V11
Fixes a bug where special characters (like apostrophes) in field names would break css selectors.
Also reordered css selector list so the confirmed working selector runs first speeding up execution time. 

Note: object id needs to be hard coded as a param in command for any non-standard objects ie. CPQ Quotes

Usage:
python main.py --file Your_Excel_File_Name.xlsx --sheet "Sheet Name" --object-api-name "API_Name__c" --object-id "validObjectId" --instance yourinstance.my.salesforce.com --start-at-label "Your Field Label Name" --limit 100
"""
import asyncio
import argparse
import sys
import openpyxl
import time
import urllib.parse
from pathlib import Path
import re

# --- Dependency Check ---
try:
    from playwright.async_api import async_playwright, TimeoutError as PWTimeout
    from bs4 import BeautifulSoup
except ImportError as e:
    missing_module_name = str(e).split("'")[1]
    install_package_name = 'beautifulsoup4' if missing_module_name == 'bs4' else missing_module_name
    print(f"--- Error: Missing Required Library ---", file=sys.stderr)
    print(f"The required Python library '{install_package_name}' is not installed.", file=sys.stderr)
    print("Please run this command in your terminal:", file=sys.stderr)
    print("\npip install \"playwright==1.44.0\" \"openpyxl==3.1.2\" \"beautifulsoup4==4.12.3\" \"lxml==5.2.2\" \"tenacity==8.3.0\"\n", file=sys.stderr)
    sys.exit(1)

# --- CLI Argument Parsing ---
parser = argparse.ArgumentParser(description="Scrape 'Where is this used?' data from Salesforce fields into Excel.")
parser.add_argument("--file", required=True, help="Path to the Excel file.")
parser.add_argument("--sheet", required=True, help="Name of the sheet containing field API names.")
parser.add_argument("--object-api-name", required=True, help="The exact Object API Name to use in the URL.")
parser.add_argument("--object-id", help="Optional: Manually provide the 15 or 18 character Object ID for custom objects.")
parser.add_argument("--start-at-label", help="The Field Label to start processing from.")
parser.add_argument("--instance", required=True, help="Your Salesforce instance domain.")
parser.add_argument("--limit", type=int, default=None, help="Maximum number of fields to process. Processes all if not specified.")
parser.add_argument("--context", default="sf_ctx", help="Directory name for browser session data.")
args = parser.parse_args()

async def navigate_to_fields_and_relationships(page, object_api_name):
    """Navigate directly to the Fields & Relationships page for our object."""
    print(f"📋 Navigating to Fields & Relationships for '{object_api_name}'...")
    
    fields_url = f"https://{args.instance}/lightning/setup/ObjectManager/{object_api_name}/FieldsAndRelationships/view"
    await page.goto(fields_url, wait_until="domcontentloaded", timeout=30000)
    await page.wait_for_timeout(3000)  # Wait for page to fully load
    
    print(f"✅ Successfully navigated to {object_api_name} Fields & Relationships")
    return fields_url # Return the URL for later use

async def extract_object_id_from_url(url):
    """Extracts the 15 or 18 character Salesforce Object ID from a setup URL."""
    # Pattern for standard and custom object IDs in an ObjectManager URL
    object_id_pattern = r'/ObjectManager/([a-zA-Z0-9]{15,18})/'
    match = re.search(object_id_pattern, url)
    if match:
        object_id = match.group(1)
        print(f"    ✅ Extracted Object ID from URL: {object_id}")
        return object_id
    print("    ⚠️ Could not extract Object ID from URL.")
    return None

async def extract_field_id_from_url(url):
    """Extract Salesforce field ID from URL."""
    try:
        # Look for field ID patterns in the URL
        
        # Pattern 1: Direct field ID in URL path (e.g., /.../FieldsAndRelationships/00N.../view)
        field_id_pattern = r'/(00N[a-zA-Z0-9]{12,15})'
        match = re.search(field_id_pattern, url)
        if match:
            field_id = match.group(1)
            print(f"    ✅ Extracted field ID from URL: {field_id}")
            return field_id
        
        # Pattern 2: Field ID in address parameter (for classic URLs wrapped in Lightning)
        if 'address=' in url:
            address_param = url.split('address=')[1].split('&')[0]
            decoded_address = urllib.parse.unquote(address_param)
            print(f"    🔍 Decoded address: {decoded_address}")
            
            match = re.search(field_id_pattern, decoded_address)
            if match:
                field_id = match.group(1)
                print(f"    ✅ Extracted field ID from address: {field_id}")
                return field_id
                
    except Exception as e:
        print(f"    ⚠️ Error extracting field ID: {e}")
    
    return None

async def navigate_to_field_dependencies_page(page, object_id, field_id):
    """Navigate directly to the field dependencies (where is this used) page."""
    try:
        # Construct direct URL to fieldDependencies page using the Object ID
        dependencies_url = f"https://{args.instance}/lightning/setup/ObjectManager/{object_id}/FieldsAndRelationships/{field_id}/fieldDependencies"
        
        print(f"    🎯 Navigating to field dependencies: {dependencies_url}")
        await page.goto(dependencies_url, wait_until="domcontentloaded", timeout=15000)
        await page.wait_for_timeout(4000)  # Wait for content to load
        
        # Check for insufficient privileges error after navigation
        if await page.locator("text=Insufficient Privileges").is_visible(timeout=1000):
            print("    ❌ Encountered 'Insufficient Privileges' error. This may be a URL format issue or a permissions problem.")
            return False

        print(f"    ✅ Successfully navigated to field dependencies page")
        return True
        
    except Exception as e:
        print(f"    ❌ Error navigating to field dependencies: {e}")
        return False

async def trigger_search_with_proper_events(page, search_box, search_term):
    """
    Trigger search with proper JavaScript events to ensure filtering works.
    """
    print(f"    🔧 Triggering search with proper events for: '{search_term}'")
    
    try:
        # Method 1: Clear, type slowly, and trigger events
        await search_box.click()
        await search_box.fill("")
        await page.wait_for_timeout(500)
        
        # Type character by character to trigger input events
        await search_box.type(search_term, delay=50)
        await page.wait_for_timeout(1000)
        
        # Trigger additional events that might be needed
        await search_box.dispatch_event("input")
        await search_box.press("Enter")
        await page.wait_for_timeout(2500) # Reduced wait time for filter
        
        print(f"    ✅ Search events triggered, waiting for table to filter...")
        
        return True
        
    except Exception as e:
        print(f"    ⚠️ Error triggering search events: {e}")
        return False

async def use_fields_page_quick_find(page, field_label, field_api_name):
    """
    Use the Quick Find box within the Fields & Relationships page to search for the field.
    Assumes page is already on the Fields & Relationships list view.
    """
    print(f"  🔍 Using Fields & Relationships Quick Find for: '{field_label}'")
    
    quick_find_selectors = [
        "input[placeholder='Quick Find']",
        "input[placeholder*='Quick Find']",
        "input[aria-label*='Quick Find']",
        ".slds-input[placeholder*='Quick Find']"
    ]
    
    search_box = None
    for selector in quick_find_selectors:
        try:
            box = page.locator(selector).first
            if await box.is_visible(timeout=5000):
                search_box = box
                print(f"    ✅ Found Fields & Relationships Quick Find: {selector}")
                break
        except:
            continue
    
    if not search_box:
        print("    ❌ Could not find Fields & Relationships Quick Find box")
        return None
    
    try:
        # Try searching by field label first with proper events
        if await trigger_search_with_proper_events(page, search_box, field_label):
            field_id = await click_field_from_table(page, field_label, field_api_name)
            if field_id:
                return field_id
        
        # If label search didn't work, try API name
        print(f"    🔍 Trying API name: {field_api_name}")
        if await trigger_search_with_proper_events(page, search_box, field_api_name):
            field_id = await click_field_from_table(page, field_label, field_api_name)
            if field_id:
                return field_id
        
        return None
        
    except Exception as e:
        print(f"    ⚠️ Fields Quick Find error: {e}")
        return None

async def click_field_from_table(page, field_label, field_api_name):
    """
    Click on the field from the filtered table results and extract the field ID.
    This version handles special characters in field labels for all selector types.
    """
    print(f"    🔍 Looking for field in filtered table...")
    
    # Robustly sanitize the label for different selector types
    escaped_for_regex = re.escape(field_label)
    escaped_for_css_string = field_label.replace("\\", "\\\\").replace("'", "\\'")

    # field_link_selectors = [
    #     f"th a:text-matches('^{escaped_for_regex}$', 'i')",
    #     f"td a:text-matches('^{escaped_for_regex}$', 'i')",
    #     f"a[title='{escaped_for_css_string}']",
    #     f"a:has-text('{escaped_for_css_string}')"
    # ]

    field_link_selectors = [
        f"a:has-text('{escaped_for_css_string}')",
        f"th a:text-matches('^{escaped_for_regex}$', 'i')",
        f"td a:text-matches('^{escaped_for_regex}$', 'i')",
        f"a[title='{escaped_for_css_string}']"
    ]
    
    for selector in field_link_selectors:
        try:
            # Wait for at least one link matching the selector to appear.
            await page.locator(selector).first.wait_for(timeout=2000)
            links = await page.locator(selector).all()
            print(f"    📊 Found {len(links)} links with selector: {selector}")
            
            for i, link in enumerate(links):
                try:
                    link_text = (await link.inner_text()).strip()
                    print(f"    📝 Checking link {i+1}: '{link_text}'")
                    
                    # Check if this link matches our field label (case-insensitive)
                    if link_text.lower() == field_label.lower():
                        print(f"    ✅ Found matching field link: {link_text}")
                        await link.click()
                        
                        await page.wait_for_load_state('domcontentloaded', timeout=10000)
                        await page.wait_for_timeout(3000)
                        
                        current_url = page.url
                        print(f"    📍 Navigated to URL: {current_url}")
                        
                        field_id = await extract_field_id_from_url(current_url)
                        if field_id:
                            return field_id
                        
                        print("    ⚠️ Could not extract field ID from URL after click.")
                        return None
                except Exception as e:
                    print(f"    ⚠️ Error checking link {i+1}: {e}")
                    continue
        except PWTimeout:
            # This is expected if a selector finds no matches, so just try the next one.
            print(f"    ... No links found with selector: {selector}")
            continue
        except Exception as e:
            print(f"    ⚠️ Error with selector '{selector}': {e}")
            continue
    
    print(f"    ❌ Field '{field_label}' not found in table after trying all selectors.")
    return None

async def find_field_and_get_dependencies(page, field_label, field_api_name, object_id):
    """
    Find field using Quick Find, get ID, and navigate to dependencies.
    Assumes the page is already on the Fields & Relationships list view.
    """
    print(f"  🎯 Finding field: '{field_label}' ({field_api_name})")
    
    field_id = await use_fields_page_quick_find(page, field_label, field_api_name)
    
    if field_id:
        if await navigate_to_field_dependencies_page(page, object_id, field_id):
            return True
    
    print(f"    ❌ Could not find field or navigate to dependencies: {field_label}")
    return False

async def find_dependencies_iframe(page):
    """Find the iframe containing the field dependencies content."""
    print("    🔍 Looking for dependencies iframe...")
    
    try:
        # More robustly wait for an iframe to be present
        await page.wait_for_selector("iframe", timeout=10000)
        iframe_count = await page.locator("iframe").count()
        print(f"    📊 Found {iframe_count} iframes")
        
        for i in range(iframe_count):
            try:
                frame = page.frame_locator(f"iframe >> nth={i}")
                
                # Check if this frame contains our table content by looking for known text
                frame_text = await frame.locator("body").inner_text(timeout=5000)
                if "Reference Type" in frame_text or "Layout" in frame_text or "Component Type" in frame_text:
                    print(f"    ✅ Found dependencies content in iframe {i}")
                    
                    # Verify the frame has our expected elements
                    await frame.locator("tr.dataRow").first.wait_for(timeout=3000)
                    return frame
            except Exception as e:
                print(f"    ⚠️ Error checking iframe {i}: {e}")
                continue
    except PWTimeout:
        print("    ❌ Timed out waiting for any iframe to appear.")

    return None

async def extract_reference_link(row, cell_index=1):
    """
    Enhanced link extraction that handles JavaScript URLs from Salesforce
    by decoding the href attribute before parsing.
    """
    try:
        cells = await row.locator(".dataCell").all()
        if len(cells) <= cell_index:
            return ""

        target_cell = cells[cell_index]
        links = await target_cell.locator("a").all()

        if not links:
            return ""

        for link in links:
            try:
                raw_href = await link.get_attribute("href")
                if not raw_href:
                    continue
                
                href = urllib.parse.unquote(raw_href)

                if href.startswith("javascript:srcUp("):
                    match = re.search(r"javascript:srcUp\('([^']+)'\)", href)
                    if match:
                        partial_url = match.group(1)
                        decoded_url = urllib.parse.unquote(partial_url)
                        if decoded_url.startswith("/"):
                            return f"https://{args.instance}{decoded_url}"
                elif href.startswith("/"):
                    return f"https://{args.instance}{href}"
                elif href.startswith("http"):
                    return href
            except Exception as e:
                print(f"      ⚠️ Error processing a link tag: {e}")
                continue
        
        return ""

    except Exception as e:
        print(f"      ⚠️ General link extraction error: {e}")
        return ""

async def parse_field_dependencies_page(page):
    """
    Parse the field dependencies page, excluding 'ReportType' and keeping the last 'Flow' instance.
    """
    field_url = page.url
    print("  📊 Parsing field dependencies page...")
    
    await page.wait_for_timeout(4000)
    
    target_frame = await find_dependencies_iframe(page)
    
    if not target_frame:
        print("    ❌ Could not find dependencies iframe. No references will be parsed.")
        return field_url, []
    
    print("    🎯 Parsing content from iframe")
    all_refs = []
    
    try:
        table = target_frame.locator("table.list").first
        
        if await table.is_visible():
            data_rows = await table.locator("tr.dataRow").all()
            print(f"    📊 Found {len(data_rows)} data rows")
            
            for i, row in enumerate(data_rows, 1):
                try:
                    cells = await row.locator(".dataCell").all()
                    if len(cells) >= 2:
                        ref_type = (await cells[0].inner_text()).strip()
                        ref_name = (await cells[1].inner_text()).strip()
                        
                        if ref_type and ref_name:
                            link_href = await extract_reference_link(row, 1)
                            all_refs.append((ref_type, ref_name, link_href))

                except Exception as e:
                    print(f"    ⚠️ Error processing row {i}: {e}")
                    continue
    except Exception as e:
        print(f"    ⚠️ Table parsing error: {e}")

    # --- NEW: Filter and deduplicate references ---
    print("    🔄 Filtering and deduplicating references...")
    filtered_refs = []
    last_flow = {} # To store the last occurrence of each flow

    for ref_type, ref_name, link in all_refs:
        # Exclude ReportType
        if ref_type == "ReportType":
            print(f"      🚫 Excluding ReportType: {ref_name}")
            continue
        
        # If it's a flow, store it. This will overwrite previous entries with the same name.
        if ref_type == "Flow":
            print(f"      🌊 Found Flow: {ref_name}. Storing as latest version.")
            last_flow[ref_name] = (ref_type, ref_name, link)
        else:
            # Add all other types directly
            filtered_refs.append((ref_type, ref_name, link))

    # Add the unique, last-seen flows back to the list
    filtered_refs.extend(last_flow.values())
    
    print(f"  ✅ Final result: {len(filtered_refs)} references after filtering.")
    return field_url, filtered_refs

async def main():
    """Main application logic."""
    if not Path(args.file).is_file():
        print(f"❌ Error: File '{args.file}' not found.")
        sys.exit(1)
    
    try:
        wb = openpyxl.load_workbook(args.file)
        ws = wb[args.sheet]
    except Exception as e:
        print(f"❌ Excel error: {e}")
        return
    
    fields_to_process = []
    start_row = 2
    
    if args.start_at_label:
        print(f"🔍 Finding start label: '{args.start_at_label}'...")
        for r_idx in range(2, ws.max_row + 1):
            label = ws.cell(row=r_idx, column=1).value
            if label and label.strip() == args.start_at_label.strip():
                start_row = r_idx
                print(f"✅ Found at row {start_row}")
                break
        else:
            print(f"❌ Start label not found")
            return
    
    limit = args.limit if args.limit is not None else (ws.max_row - start_row + 1)
    end_row = start_row + limit
    for row_idx in range(start_row, min(end_row, ws.max_row + 1)):
        api_name_cell = ws.cell(row=row_idx, column=2)
        label_cell = ws.cell(row=row_idx, column=1)
        
        # Only process rows that have a label and haven't been processed yet
        if api_name_cell.value and label_cell.value and not ws.cell(row=row_idx + 1, column=9).value:
            fields_to_process.append({
                "api_name": api_name_cell.value,
                "label": label_cell.value,
                "original_row": row_idx
            })
    
    if not fields_to_process:
        print("✅ No more fields to process or all remaining fields have been processed.")
        return
    
    print(f"🚀 Processing {len(fields_to_process)} fields...")
    
    async with async_playwright() as p:
        try:
            browser = await p.chromium.launch_persistent_context(
                args.context,
                headless=False,
                viewport={'width': 1600, 'height': 1200},
                slow_mo=100
            )
            page = browser.pages[0] if browser.pages else await browser.new_page()
            
            print(f"🌐 Navigating to Salesforce...")
            await page.goto(f"https://{args.instance}", timeout=30000)
            
            try:
                await page.wait_for_selector("div.slds-icon-waffle", timeout=10000)
                print("✅ Session found")
            except PWTimeout:
                print("\n🔐 Please log in to Salesforce and press Enter when ready...")
                input()
                await page.wait_for_selector("div.slds-icon-waffle", timeout=60000)
                print("✅ Login confirmed")
            
            # --- UPDATED: Navigate once and extract the Object ID ---
            print("\n--- Performing initial navigation to Fields & Relationships page ---")
            fields_list_url = await navigate_to_fields_and_relationships(page, args.object_api_name)
            
            object_id = None
            if args.object_id:
                print(f"🔵 Using manually provided Object ID: {args.object_id}")
                object_id = args.object_id
            else:
                print("    🔍 Attempting to extract Object ID from URL...")
                canonical_url = page.url
                object_id = await extract_object_id_from_url(canonical_url)

            if not object_id:
                print(f"❌ CRITICAL: Could not determine Object ID for '{args.object_api_name}'. Please provide it using --object-id. Exiting.")
                return

            rows_inserted = 0
            
            for i, field in enumerate(fields_to_process, 1):
                current_row = field["original_row"] + rows_inserted
                field_label = field["label"]
                field_api_name = field["api_name"]
                
                print(f"\n--- Field {i}/{len(fields_to_process)}: {field_label} (Sheet Row {current_row}) ---")
                
                try:
                    # Pass the extracted object_id to the function chain
                    if await find_field_and_get_dependencies(page, field_label, field_api_name, object_id):
                        field_url, refs = await parse_field_dependencies_page(page)
                        
                        print("    🔄 Navigating back to Fields & Relationships for next search...")
                        await page.goto(fields_list_url, wait_until="domcontentloaded")
                        await page.wait_for_timeout(3000)


                        field_cell = ws.cell(row=current_row, column=1)
                        field_cell.hyperlink = field_url
                        
                        if refs:
                            print(f"    📊 Adding {len(refs)} references to Excel")
                            ws.insert_rows(current_row + 1, amount=len(refs))
                            for j, (ref_type, ref_name, link) in enumerate(refs):
                                insert_row_index = current_row + 1 + j
                                ws.cell(row=insert_row_index, column=9, value=ref_type)
                                ref_cell = ws.cell(row=insert_row_index, column=10, value=ref_name)
                                if link:
                                    ref_cell.hyperlink = link
                            rows_inserted += len(refs)
                        else:
                            # If no refs found after filtering, mark it as processed by adding a note
                            ws.insert_rows(current_row + 1, amount=1)
                            ws.cell(row=current_row + 1, column=9, value="No references found")
                            rows_inserted += 1
                            print("    📊 No references found for this field")
                    else:
                        print(f"    ❌ Skipped: {field_label}. Navigating back to list.")
                        await page.goto(fields_list_url, wait_until="domcontentloaded")
                        await page.wait_for_timeout(3000)
                except Exception as e:
                    print(f"    ❌ An error occurred while processing {field_label}: {e}")
                    import traceback
                    traceback.print_exc()
                    try:
                        print("    🔄 Attempting to recover by navigating back to Fields & Relationships...")
                        await page.goto(fields_list_url, wait_until="domcontentloaded")
                        await page.wait_for_timeout(3000)
                    except Exception as nav_e:
                        print(f"    ❌ CRITICAL: Failed to recover by navigating. Stopping script. Error: {nav_e}")
                        break
                    continue
        except Exception as e:
            print(f"❌ A critical browser error occurred: {e}")
            import traceback
            traceback.print_exc()
        finally:
            print("\n--- Script finished. Saving and closing browser. ---")
            if 'browser' in locals() and browser:
                await browser.close()
    
    try:
        save_path = Path(args.file)
        wb.save(save_path)
        print(f"\n✅ Results successfully saved to {save_path.resolve()}")
    except Exception as e:
        print(f"\n❌ Error saving Excel file: {e}. Please ensure it is not open.")

if __name__ == "__main__":
    asyncio.run(main())